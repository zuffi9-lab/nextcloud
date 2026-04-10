#!/usr/bin/env python3
import argparse
import datetime as dt
import hashlib
import io
import json
import sqlite3
from dataclasses import dataclass
from typing import Dict, Optional

import openpyxl
import requests
from dateutil import parser as date_parser


@dataclass
class Config:
    raw: Dict

    @property
    def nextcloud(self):
        return self.raw["nextcloud"]

    @property
    def telegram(self):
        return self.raw["telegram"]

    @property
    def database_path(self):
        return self.raw.get("database", {}).get("path", "state.db")

    @property
    def columns(self):
        return self.raw.get("columns", {})

    @property
    def timezone(self):
        return self.raw.get("event", {}).get("timezone", "Europe/Moscow")


def read_config(path: str) -> Config:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return Config(raw=data)


def ensure_db(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS reminders (
            key TEXT PRIMARY KEY,
            item_id TEXT NOT NULL,
            kind TEXT NOT NULL,
            last_date TEXT NOT NULL,
            last_notified_on TEXT,
            etag TEXT
        )
        """
    )
    conn.commit()


def webdav_download_xlsx(cfg: Config) -> bytes:
    nc = cfg.nextcloud
    url = f"{nc['base_url'].rstrip('/')}{nc['xlsx_webdav_path']}"
    resp = requests.get(url, auth=(nc["username"], nc["app_password"]), timeout=60)
    resp.raise_for_status()
    return resp.content


def parse_date(value) -> Optional[dt.date]:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)):
        return None
    parsed = date_parser.parse(str(value), dayfirst=True)
    return parsed.date()


def load_rows(xlsx_bytes: bytes, cfg: Config):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    c_id = cfg.columns.get("id", "id")
    c_title = cfg.columns.get("title", "title")
    c_fn = cfg.columns.get("fn_expiry_date", "fn_expiry_date")
    c_ofd = cfg.columns.get("ofd_expiry_date", "ofd_expiry_date")

    for required in [c_id, c_title, c_fn, c_ofd]:
        if required not in idx:
            raise ValueError(f"В таблице не найдена колонка: {required}")

    items = []
    for r in ws.iter_rows(min_row=2):
        rid = r[idx[c_id]].value
        if rid is None or str(rid).strip() == "":
            continue
        title = r[idx[c_title]].value
        fn_date = parse_date(r[idx[c_fn]].value)
        ofd_date = parse_date(r[idx[c_ofd]].value)
        items.append(
            {
                "id": str(rid).strip(),
                "title": str(title).strip() if title else str(rid).strip(),
                "fn_expiry_date": fn_date,
                "ofd_expiry_date": ofd_date,
            }
        )
    return items


def reminder_key(item_id: str, kind: str) -> str:
    return f"{item_id}:{kind}"


def get_state(conn: sqlite3.Connection, key: str):
    row = conn.execute(
        "SELECT item_id, kind, last_date, last_notified_on, etag FROM reminders WHERE key=?", (key,)
    ).fetchone()
    if not row:
        return None
    return {
        "item_id": row[0],
        "kind": row[1],
        "last_date": row[2],
        "last_notified_on": row[3],
        "etag": row[4],
    }


def upsert_state(conn: sqlite3.Connection, key: str, item_id: str, kind: str, last_date: dt.date,
                 last_notified_on: Optional[dt.date] = None, etag: Optional[str] = None):
    conn.execute(
        """
        INSERT INTO reminders(key, item_id, kind, last_date, last_notified_on, etag)
        VALUES(?,?,?,?,?,?)
        ON CONFLICT(key) DO UPDATE SET
          last_date=excluded.last_date,
          last_notified_on=COALESCE(excluded.last_notified_on, reminders.last_notified_on),
          etag=COALESCE(excluded.etag, reminders.etag)
        """,
        (
            key,
            item_id,
            kind,
            last_date.isoformat(),
            last_notified_on.isoformat() if last_notified_on else None,
            etag,
        ),
    )
    conn.commit()


def should_notify(today: dt.date, expiry: dt.date, last_notified_on: Optional[str]) -> bool:
    days_left = (expiry - today).days
    if last_notified_on:
        last = dt.date.fromisoformat(last_notified_on)
        since = (today - last).days
    else:
        since = 9999

    if days_left == 30:
        return True
    if days_left < 30 and days_left >= 0 and since >= 10:
        return True
    if days_left < 0 and since >= 10:
        return True
    return False


def compose_message(item_title: str, kind: str, expiry: dt.date, today: dt.date) -> str:
    days_left = (expiry - today).days
    label = "ФН" if kind == "fn" else "ОФД"
    if days_left > 0:
        return (
            f"⏰ {item_title}: срок {label} истекает {expiry.isoformat()} "
            f"(осталось {days_left} дн.)."
        )
    if days_left == 0:
        return f"⚠️ {item_title}: срок {label} истекает сегодня ({expiry.isoformat()})."
    return (
        f"❗ {item_title}: срок {label} истёк {expiry.isoformat()} "
        f"({abs(days_left)} дн. назад). Обновите данные в таблице Nextcloud."
    )


def send_telegram(cfg: Config, text: str) -> None:
    tg = cfg.telegram
    url = f"https://api.telegram.org/bot{tg['bot_token']}/sendMessage"
    resp = requests.post(url, json={"chat_id": tg["chat_id"], "text": text}, timeout=30)
    resp.raise_for_status()


def build_ics(uid: str, summary: str, event_date: dt.date, description: str) -> str:
    d = event_date.strftime("%Y%m%d")
    created = dt.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    return "\r\n".join(
        [
            "BEGIN:VCALENDAR",
            "VERSION:2.0",
            "PRODID:-//nextcloud-fn-ofd-reminder//EN",
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{created}",
            f"DTSTART;VALUE=DATE:{d}",
            f"DTEND;VALUE=DATE:{(event_date + dt.timedelta(days=1)).strftime('%Y%m%d')}",
            f"SUMMARY:{summary}",
            f"DESCRIPTION:{description}",
            "END:VEVENT",
            "END:VCALENDAR",
            "",
        ]
    )


def upsert_calendar_event(cfg: Config, key: str, summary: str, event_date: dt.date, description: str):
    cal_url = cfg.nextcloud["calendar_url"].rstrip("/") + "/"
    uid = hashlib.md5(key.encode("utf-8")).hexdigest() + "@fn-ofd"
    href = f"{uid}.ics"
    ics = build_ics(uid=uid, summary=summary, event_date=event_date, description=description)
    url = cal_url + href
    resp = requests.put(
        url,
        data=ics.encode("utf-8"),
        auth=(cfg.nextcloud["username"], cfg.nextcloud["app_password"]),
        headers={"Content-Type": "text/calendar; charset=utf-8"},
        timeout=60,
    )
    resp.raise_for_status()
    return resp.headers.get("ETag")


def process(cfg: Config):
    today = dt.date.today()
    conn = sqlite3.connect(cfg.database_path)
    ensure_db(conn)

    xlsx = webdav_download_xlsx(cfg)
    items = load_rows(xlsx, cfg)

    for item in items:
        for kind, date_field in [("fn", "fn_expiry_date"), ("ofd", "ofd_expiry_date")]:
            expiry = item[date_field]
            if not expiry:
                continue

            key = reminder_key(item["id"], kind)
            st = get_state(conn, key)

            if st is None:
                upsert_state(conn, key, item["id"], kind, expiry)
                st = get_state(conn, key)

            prev_date = dt.date.fromisoformat(st["last_date"])
            if expiry != prev_date:
                upsert_state(conn, key, item["id"], kind, expiry, last_notified_on=None)
                st = get_state(conn, key)

            summary = f"{item['title']} — замена {('ФН' if kind == 'fn' else 'ОФД')}"
            description = f"Автособытие. Истечение срока {kind.upper()} для {item['title']}"
            etag = upsert_calendar_event(cfg, key, summary, expiry, description)
            upsert_state(conn, key, item["id"], kind, expiry, etag=etag)

            if should_notify(today, expiry, st["last_notified_on"]):
                msg = compose_message(item["title"], kind, expiry, today)
                send_telegram(cfg, msg)
                upsert_state(conn, key, item["id"], kind, expiry, last_notified_on=today, etag=etag)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="FN/OFD reminder for Nextcloud + Telegram")
    parser.add_argument("--config", default="config.json", help="Path to config JSON")
    args = parser.parse_args()

    config = read_config(args.config)
    process(config)
